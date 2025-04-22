import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk, Canvas
import csv
import requests
import threading
from datetime import datetime
import pandas as pd
import webbrowser
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from nltk.sentiment import SentimentIntensityAnalyzer
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# --- Tooltip Class ---
class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind("<Enter>", self.show)
        widget.bind("<Leave>", self.hide)

    def show(self, event=None):
        if self.tip_window:
            return
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, background="#ffffe0", relief="solid", borderwidth=1, font=("tahoma", "10", "normal"))
        label.pack(ipadx=1)

    def hide(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None

# --- Reddit Analyzer ---
class RedditAnalyzerGUI:
    def __init__(self, master):
        self.master = master
        master.title("RedditWale Baba - Profile Analyzer")
        master.geometry("1000x700")  # Set window size

        self.username_list = []
        self.user_metadata = {}

        # Create a notebook for tabs
        self.notebook = ttk.Notebook(master)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Create frames for each tab
        self.text_frame = tk.Frame(self.notebook)
        self.chart_frame = tk.Frame(self.notebook)
        self.help_frame = tk.Frame(self.notebook)

        self.notebook.add(self.text_frame, text="Text Output")
        self.notebook.add(self.chart_frame, text="Summary Chart")
        self.notebook.add(self.help_frame, text="Help / Manual")

        # Text Output Tab
        self.output_canvas = Canvas(self.text_frame)
        self.output_canvas.pack(fill=tk.BOTH, expand=True)

        # Bind the canvas to the <Configure> event to redraw the gradient on resize
        self.output_canvas.bind("<Configure>", self.redraw_gradient)

        # Create a card-like frame for text output
        self.output_card = tk.Frame(self.output_canvas, bg="white", bd=2, relief="groove")
        self.output_card.place(relx=0.5, rely=0.4, anchor="center", width=900, height=400)

        # Add a scrolled text widget inside the card
        self.output_text = scrolledtext.ScrolledText(self.output_card, wrap=tk.WORD, font=("Courier", 10), bg="white", bd=0)
        self.output_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        # Buttons Frame (moved inside text_frame)
        self.btn_frame = tk.Frame(self.text_frame, bg="white")
        self.btn_frame.place(relx=0.5, rely=0.8, anchor="center")

        # Buttons
        self.btn1 = tk.Button(self.btn_frame, text="Select CSV Only with Usernames", command=self.load_csv, width=25)
        self.btn1.grid(row=0, column=0, padx=10, pady=5)
        ToolTip(self.btn1, "Select a CSV file containing Reddit usernames.")

        self.btn2 = tk.Button(self.btn_frame, text="Track", command=self.analyze_profiles, width=15)
        self.btn2.grid(row=0, column=1, padx=10, pady=5)
        ToolTip(self.btn2, "Start analyzing the loaded usernames.")

        self.export_button = tk.Button(self.btn_frame, text="Export to Excel", command=self.export_to_excel, width=20)
        self.export_button.grid(row=0, column=2, padx=10, pady=5)
        ToolTip(self.export_button, "Export the analysis results to an Excel file.")

        # Progress Label
        self.progress_label = tk.Label(self.text_frame, text="", font=("Arial", 10), bg="white")
        self.progress_label.place(relx=0.5, rely=0.9, anchor="center")

        # Help Tab
        self.help_text = tk.Text(self.help_frame, wrap=tk.WORD, font=("Arial", 10))
        self.help_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        self.help_text.insert(tk.END, """
🔍 How to Use RedditWale Baba:

1. Click 'Select CSV Only with Usernames' and choose a CSV file with usernames.
2. Click 'Track' to analyze profiles.
3. Use the 'Export to Excel' button to save results.
4. Use the 'Summary Chart' tab to visualize karma.
5. All links in the output are clickable.

🔧 Tips:
- Ensure internet connectivity.
- Each Reddit username should be on a new line in the CSV.

📞 Contact: support@redditwalebaba.com
""")
        self.help_text.config(state='disabled')

    def draw_gradient(self, canvas, color1, color2, width, height):
        """Draw a vertical gradient on the canvas."""
        limit = 256  # Number of gradient steps
        r1, g1, b1 = self.hex_to_rgb(color1)
        r2, g2, b2 = self.hex_to_rgb(color2)

        for i in range(limit):
            r = int(r1 + (r2 - r1) * (i / limit))
            g = int(g1 + (g2 - g1) * (i / limit))
            b = int(b1 + (b2 - b1) * (i / limit))
            color = f"#{r:02x}{g:02x}{b:02x}"
            canvas.create_rectangle(0, i * height / limit, width, (i + 1) * height / limit, outline="", fill=color)

    def redraw_gradient(self, event):
        """Redraw the gradient when the canvas is resized."""
        self.draw_gradient(self.output_canvas, "#FFDEE9", "#B5FFFC", event.width, event.height)

    def hex_to_rgb(self, hex_color):
        """Convert a hex color to an RGB tuple."""
        hex_color = hex_color.lstrip("#")
        return tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))

    def load_csv(self):
        file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if not file_path:
            return
        with open(file_path, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            self.username_list = [row[0].strip() for row in reader if row]
        messagebox.showinfo("CSV Loaded", f"Loaded {len(self.username_list)} usernames.")

    def analyze_profiles(self):
        if not self.username_list:
            messagebox.showwarning("No Data", "Please load a CSV with usernames first.")
            return
        threading.Thread(target=self.thread_task).start()

    def thread_task(self):
        self.output_text.delete("1.0", tk.END)
        self.user_metadata = {}
        total = len(self.username_list)
        sia = SentimentIntensityAnalyzer()  # Initialize sentiment analyzer

        for idx, username in enumerate(self.username_list, 1):
            url = f"https://www.reddit.com/user/{username}/about.json"
            headers = {'User-Agent': 'Mozilla/5.0'}
            try:
                res = requests.get(url, headers=headers, timeout=10)
                if res.status_code == 200:
                    data = res.json().get('data', {})
                    name = data.get('name', 'N/A')
                    link = f"https://www.reddit.com/user/{name}"
                    post_karma = data.get('link_karma', 0)  # Fetch post karma
                    comment_karma = data.get('comment_karma', 0)  # Fetch comment karma
                    total_karma = post_karma + comment_karma  # Total karma (post + comment)
                    created = datetime.utcfromtimestamp(data.get('created_utc', 0)).strftime('%Y-%m-%d')
                    account_age_days = (datetime.utcnow() - datetime.utcfromtimestamp(data.get('created_utc', 0))).days
                    karma_per_day = total_karma

                    # Fetch posts and comments
                    posts_url = f"https://www.reddit.com/user/{username}/submitted.json"
                    comments_url = f"https://www.reddit.com/user/{username}/comments.json"
                    posts = self.fetch_reddit_data(posts_url, headers)
                    comments = self.fetch_reddit_data(comments_url, headers)

                    # Process posts and comments
                    items = []
                    items.extend(self.process_items(posts, "P", sia, karma_per_day))
                    items.extend(self.process_items(comments, "C", sia, karma_per_day))

                    self.user_metadata[name] = {
                        "karma": total_karma,
                        "created": created,
                        "account_age_days": account_age_days,
                        "karma_per_day": karma_per_day,
                        "items": items
                    }

                    self.output_text.insert(tk.END, f"{name} ({created})\n")
                    self.output_text.insert(tk.END, f"Link: {link}\n", ("link",))
                    self.output_text.insert(tk.END, f"Karma: {total_karma} | Karma/Day: {karma_per_day}\n\n")
                else:
                    # Corrected failure message
                    self.output_text.insert(tk.END, f"{username} :- failed to retrieve\n\n")
            except Exception as e:
                # Corrected failure message
                self.output_text.insert(tk.END, f"{username} :- failed to retrieve\n\n")
            self.progress_label.config(text=f"Analyzing {idx}/{total}...")

        self.output_text.tag_config("link", foreground="blue", underline=True)
        self.output_text.tag_bind("link", "<Button-1>", self.open_link)
        self.progress_label.config(text="Done.")
        self.show_summary_chart()

    def fetch_reddit_data(self, url, headers):
        try:
            res = requests.get(url, headers=headers, timeout=10)
            if res.status_code == 200:
                return res.json().get('data', {}).get('children', [])
        except Exception as e:
            print(f"Error fetching data: {e}")
        return []

    def process_items(self, items, item_type, sia, karma_per_day):
        processed_items = []
        for item in items:
            data = item.get('data', {})
            title_or_body = data.get('title', '') if item_type == "P" else data.get('body', '')
            sentiment = sia.polarity_scores(title_or_body)
            positive_count = 1 if sentiment['compound'] > 0.05 else 0
            negative_count = 1 if sentiment['compound'] < -0.05 else 0

            processed_items.append({
                "Sno": len(processed_items) + 1,
                "Type": item_type,
                "Name": data.get('title', 'N/A') if item_type == "P" else title_or_body[:100],
                "Account": data.get('author', 'N/A'),
                "Account-Link": f"https://www.reddit.com/user/{data.get('author', '')}",
                "Subreddit": data.get('subreddit', 'N/A'),
                "Subreddit(O/G)": "G",
                "Comment/Post-Link": f"https://www.reddit.com{data.get('permalink', '')}",
                "Related-To-LPU": "Yes" if "lpu" in title_or_body.lower() else "No",
                "Views": "N/A",
                "Post-Upvote-Count": data.get('ups', 0),
                "No-of-comments-in-the-post": data.get('num_comments', 0) if item_type == "P" else "N/A",
                "Positive-Comments-Count": positive_count,
                "Negative-Comments-Count": negative_count,
                "No-of-Comments-": data.get('num_comments', 0) if item_type == "P" else "N/A",
                "Comment-Upvote-Count": data.get('ups', 0) if item_type == "C" else "N/A",
                "Karma-Points-/-Day": karma_per_day
            })
        return processed_items

    def show_summary_chart(self):
        for widget in self.chart_frame.winfo_children():
            widget.destroy()

        usernames = list(self.user_metadata.keys())
        karma = [self.user_metadata[u]["karma"] for u in usernames]

        fig = Figure(figsize=(6, 4), dpi=100)
        ax = fig.add_subplot(111)
        ax.bar(usernames, karma, color="skyblue")
        ax.set_title("Total Karma by User")
        ax.set_ylabel("Karma")
        ax.set_xticklabels(usernames, rotation=45, ha="right")

        canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def export_to_excel(self):
        if not self.user_metadata:
            messagebox.showwarning("No Data", "Nothing to export.")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if not file_path:
            return

        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Reddit Analysis Report"

        # Define headers
        headers = [
            "Sno", "Type", "Title", "Account", "Account-Link", "Subreddit",
            "Subreddit(O/G)", "Comment/Post-Link", "Related-To-LPU", "Views",
            "Post-Upvote-Count", "No-of-comments-in-the-post",
            "Positive-Comments-Count", "Negative-Comments-Count", "No-of-Comments-",
            "Comment-Upvote-Count", "Karma-Points-/-Day"
        ]

        # Add headers to the worksheet
        ws.append(headers)

        # Style the header row (yellow fill and bold text)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:  # First row contains the headers
            cell.fill = header_fill
            cell.font = header_font

        # Add data rows
        rows = []
        for user, metadata in self.user_metadata.items():
            rows.extend(metadata["items"])

        for row in rows:
            ws.append([
                row["Sno"], row["Type"], row["Name"], row["Account"], row["Account-Link"],
                row["Subreddit"], row["Subreddit(O/G)"], row["Comment/Post-Link"],
                row["Related-To-LPU"], row["Views"], row["Post-Upvote-Count"],
                row["No-of-comments-in-the-post"], row["Positive-Comments-Count"],
                row["Negative-Comments-Count"], row["No-of-Comments-"],
                row["Comment-Upvote-Count"], row["Karma-Points-/-Day"]
            ])

        # Save the workbook
        wb.save(file_path)
        messagebox.showinfo("Exported", f"Data exported to {file_path}")

    def open_link(self, event):
        
        index = self.output_text.index("@%s,%s" % (event.x, event.y))
        line = self.output_text.get(index + " linestart", index + " lineend")
        if line.startswith("Link: "):
            url = line.split("Link: ")[1]
            webbrowser.open_new(url)

if __name__ == "__main__":
    root = tk.Tk()
    app = RedditAnalyzerGUI(root)
    root.mainloop()
